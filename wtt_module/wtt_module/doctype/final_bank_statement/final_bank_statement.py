# Copyright (c) 2022, wtt_module and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.model.mapper import get_mapped_doc
from num2words import num2words
from datetime import datetime,date,timedelta
from frappe.utils.xlsxutils import make_xlsx
from six import string_types
import json
import requests
import openpyxl
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

class Finalbankstatement(Document):
    def validate(self):
        vv=str(self.posting_date)
        vs=datetime.strptime(vv,"%Y-%m-%d")
        d=vs
        v=str(d.strftime('%d'))+""+d.strftime('%m')+""+d.strftime('%Y')
        imp = '&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;'.join([str(v[i]) for i in range(len(v))])
        self.check_date_format=imp
    def on_update(self):
        if self.approved_payment:
            self.amount_table=[]
            for i in frappe.db.sql("SELECT distinct(benificiary_name)as benificiary_name,party_type,bank_name,account_no,branch,ifsc_code,module_name,sum(amount)as amount,GROUP_CONCAT(distinct(request_no))as request_no,line FROM `tabFBS Table` WHERE parent='"+str(self.name)+"' GROUP BY benificiary_name ",as_dict=1):
                self.append("amount_table",i)
        if self.approved_payment_ib:
            self.ib_amount_table=[]
            for i in frappe.db.sql("SELECT distinct(benificiary_name)as benificiary_name,party_type,bank_name,account_no,branch,ifsc_code,module_name,sum(amount)as amount,GROUP_CONCAT(distinct(request_no))as request_no,line FROM `tabFBS Table IB` WHERE parent='"+str(self.name)+"' GROUP BY benificiary_name ",as_dict=1):
                self.append("ib_amount_table",i)


        valword=num2words(self.net_total, lang='en_IN')
        self.net_total2=valword.capitalize()
        ib_valword=num2words(self.ib_net_total, lang='en_IN')
        self.ib_net_total2=ib_valword.capitalize()

    def on_cancel(self):
        frappe.db.sql("DELETE FROM `tabPayment Entry` WHERE final_bank_statement='"+str(self.name)+"' and docstatus=0 ")

    
    def on_submit(self):
        supls=[]
        if(self.amount_table):
            for i in frappe.db.sql("SELECT distinct(benificiary_name)as benificiary_name,party_type,sum(amount)as amount FROM `tabBank Statement Table` WHERE parent='"+str(self.name)+"'  GROUP BY benificiary_name ",as_dict=1):
                # if(i.benificiary_name!='MANJEET COTTON PRIVATE LIMITED'):
                supls.append({
                    "supplier":i.benificiary_name,
                    "amount":i.amount,
                    "party_type":i.party_type,
                    "bank":"OTHER BANK"
                    })
        if(self.ib_amount_table):
            for i in frappe.db.sql("SELECT distinct(benificiary_name)as benificiary_name,party_type,sum(amount)as amount FROM `tabBank Statement Table IB` WHERE parent='"+str(self.name)+"'  GROUP BY benificiary_name ",as_dict=1):
                supls.append({
                    "supplier":i.benificiary_name,
                    "amount":i.amount,
                    "party_type":i.party_type,
                    "bank":"INDIAN BANK"
                    })


        rfq_suppliers=[]
        if(self.statement_advance_table):
            for j in supls:
                ar=[]
                sub_query = frappe.db.sql("SELECT distinct(order_no)as order_no,cheque_no,supplier,final_amount FROM `tabStatement Advance Table` WHERE parent='"+str(self.name)+"' and supplier='"+str(j["supplier"])+"' ",as_dict=1)
                if sub_query:
                    for i in sub_query:
                        rfq_suppliers.append(i.supplier)
                    
                    doc=frappe.new_doc("Payment Entry")
                    doc.payment_type='Pay'
                    doc.mode_of_payment='NEFT/RTGS/IMPS'
                    doc.party_type='Supplier'
                    doc.party=str(j["supplier"])
                    doc.paid_from=self.paid_from
                    doc.paid_amount=float(j["amount"])
                    doc.received_amount=float(j["amount"])
                    doc.total_allocated_amount=float(j["amount"])
                    doc.source_exchange_rate=1
                    doc.target_exchange_rate=1
                    doc.final_bank_statement=self.name
                    for i in sub_query:
                        doc.append("references",{
                            "reference_doctype":"Purchase Order",
                            "reference_name":i.order_no,
                            "allocated_amount":float(i.final_amount)
                            })
                        doc.reference_no=i.cheque_no

                    doc.reference_date=date.today()
                    doc.save()


        if(self.statement_freight_table):
            for j in supls:
                ar=[]
                sub_query = frappe.db.sql("SELECT distinct(order_no)as order_no,cheque_no,supplier,final_amount FROM `tabStatement Freight Table` WHERE parent='"+str(self.name)+"' and supplier='"+str(j["supplier"])+"' ",as_dict=1)
                if sub_query:
                    for i in sub_query:
                        rfq_suppliers.append(i.supplier)

                    doc=frappe.new_doc("Payment Entry")
                    doc.payment_type='Pay'
                    doc.mode_of_payment='NEFT/RTGS/IMPS'
                    doc.party_type='Supplier'
                    doc.party=str(j["supplier"])
                    doc.paid_from=self.paid_from
                    doc.paid_amount=float(j["amount"])
                    doc.received_amount=float(j["amount"])
                    doc.total_allocated_amount=float(j["amount"])
                    doc.source_exchange_rate=1
                    doc.target_exchange_rate=1
                    doc.final_bank_statement=self.name
                    for i in sub_query:
                        # doc.append("references",{
                        #     "reference_doctype":"Purchase Order",
                        #     "reference_name":i.order_no,
                        #     "allocated_amount":float(i.final_amount)
                        #     })
                        doc.reference_no=i.cheque_no
                    doc.reference_date=date.today()
                    doc.save()

        if(self.statement_overdue_table):
            for j in supls:
                ar=[]
                sub_query = frappe.db.sql("SELECT distinct(invoice_no)as order_no,cheque_no,supplier,outstanding_amount FROM `tabStatement Overdue Table` WHERE parent='"+str(self.name)+"' and supplier='"+str(j["supplier"])+"' ",as_dict=1)
                if sub_query:
                    for i in sub_query:
                        rfq_suppliers.append(i.supplier)

                    doc=frappe.new_doc("Payment Entry")
                    doc.payment_type='Pay'
                    doc.mode_of_payment='NEFT/RTGS/IMPS'
                    doc.party_type='Supplier'
                    doc.party=str(j["supplier"])
                    doc.paid_from=self.paid_from
                    doc.paid_amount=float(j["amount"])
                    doc.received_amount=float(j["amount"])
                    doc.total_allocated_amount=float(j["amount"])
                    doc.source_exchange_rate=1
                    doc.target_exchange_rate=1
                    doc.final_bank_statement=self.name
                    for i in sub_query:
                        doc.append("references",{
                            "reference_doctype":"Purchase Invoice",
                            "reference_name":i.order_no,
                            "allocated_amount":i.outstanding_amount
                            })
                        doc.reference_no=i.cheque_no
                    doc.reference_date=date.today()
                    doc.save()

        for i in supls:
            if(i["supplier"] not in rfq_suppliers):
                doc=frappe.new_doc("Payment Entry")
                doc.payment_type='Pay'
                doc.mode_of_payment='NEFT/RTGS/IMPS'
                doc.party_type=i["party_type"]
                doc.party=str(i["supplier"])
                doc.paid_from=self.paid_from
                doc.paid_amount=float(i["amount"])
                doc.total_allocated_amount=float(i["amount"])
                doc.source_exchange_rate=1
                doc.target_exchange_rate=1
                doc.final_bank_statement=self.name
                doc.reference_no=self.cheque
                if(i["bank"]=="INDIAN BANK"):
                    doc.reference_no=self.ib_cheque
                doc.reference_date=date.today()
                doc.paid_to=self.paid_from
                doc.received_amount=float(i["amount"])
                doc.save()

                #frappe.msgprint("Kindly choose the Paid to for "+str(i["party_name"])+" Submit the Payment Entry "+doc.name)
                # doc.paid_to=None
                # doc.received_amount=float(i["amount"])
                # doc.submit()



    @frappe.whitelist()
    def num_to_word(self):
        valword=num2words(self.net_total, lang='en_IN')
        self.net_total2=valword.capitalize()
        ib_valword=num2words(self.ib_net_total, lang='en_IN')
        self.ib_net_total2=ib_valword.capitalize()
        
@frappe.whitelist()
def make_advance(source_name, target_doc=None):
    def postprocess(source, target):
        item=[]
        for i in source.get('advance_consolidate_table'):
            for j in frappe.db.sql("SELECT bank_account_no,ifsc_code,bank,branch FROM `tabBank Account` WHERE party='"+i.supplier+"'",as_dict=1):
                if(float(i.advance_amount)<=float(10000)):
                    crg=2
                elif(float(i.advance_amount)>float(10000) and float(i.advance_amount)<=float(100000)):
                    crg=5
                elif(float(i.advance_amount)>float(100000) and float(i.advance_amount)<=float(200000)):
                    crg=13
                elif(float(i.advance_amount)>float(200000) and float(i.advance_amount)<=float(500000)):
                    crg=28
                elif(float(i.advance_amount)>float(500000)):
                    crg=56
                item.append({
                    "beneficiary_name":i.supplier,
                    "account_no":j.bank_account_no,
                    "ifsc_code":j.ifsc_code,
                    "bank_name":j.bank,
                    "branch":j.branch,
                    "charges":crg,
                    "amount":i.advance_amount,
                    "request_no":i.name
                    })
        for i in item:
            if(i["bank_name"]=='INDIAN BANK'):
                target.append("ib_amount_table",{
                    "benificiary_name":i["beneficiary_name"],
                    "party_type":"Supplier",
                    "party_name":i["beneficiary_name"],
                    "account_no":i["account_no"],
                    "ifsc_code":i["ifsc_code"],
                    "bank_name":i["bank_name"],
                    "branch":i["branch"],
                    "charges":0,
                    "amount":i["amount"],
                    "request_no":i["request_no"],
                    "module_name":"Advance Payment"
                    })
            else:
                target.append("amount_table",{
                    "benificiary_name":i["beneficiary_name"],
                    "party_type":"Supplier",
                    "party_name":i["beneficiary_name"],
                    "account_no":i["account_no"],
                    "ifsc_code":i["ifsc_code"],
                    "bank_name":i["bank_name"],
                    "branch":i["branch"],
                    "charges":i["charges"],
                    "amount":i["amount"],
                    "request_no":i["request_no"],
                    "module_name":"Advance Payment"
                    })
    doc = get_mapped_doc("Payment Module", source_name, {
        "Payment Module": {
            "doctype": "Final bank statement",
            "validation": {
                "docstatus": ["=", 1]
            }
        },
        "Advance Table":{
            "doctype":"Statement Advance Table"
        }
    }, target_doc,postprocess)
    return doc

@frappe.whitelist()
def make_freight(source_name, target_doc=None):
    def postprocess(source, target):
        item=[]
        for i in source.get('freight_consolidate_table'):
            for j in frappe.db.sql("SELECT bank_account_no,ifsc_code,bank,branch FROM `tabBank Account` WHERE party='"+i.supplier+"'",as_dict=1):
                if(float(i.advance_amount)<=float(10000)):
                    crg=2
                elif(float(i.advance_amount)>float(10000) and float(i.advance_amount)<=float(100000)):
                    crg=5
                elif(float(i.advance_amount)>float(100000) and float(i.advance_amount)<=float(200000)):
                    crg=13
                elif(float(i.advance_amount)>float(200000) and float(i.advance_amount)<=float(500000)):
                    crg=28
                elif(float(i.advance_amount)>float(500000)):
                    crg=56
                item.append({
                    "beneficiary_name":i.supplier,
                    "account_no":j.bank_account_no,
                    "ifsc_code":j.ifsc_code,
                    "bank_name":j.bank,
                    "branch":j.branch,
                    "charges":crg,
                    "amount":i.advance_amount,
                    "request_no":i.name
                    })
        for i in item:
            if(i["bank_name"]=='INDIAN BANK'):
                target.append("ib_amount_table",{
                    "benificiary_name":i["beneficiary_name"],
                    "party_type":"Supplier",
                    "party_name":i["beneficiary_name"],
                    "account_no":i["account_no"],
                    "ifsc_code":i["ifsc_code"],
                    "bank_name":i["bank_name"],
                    "branch":i["branch"],
                    "charges":0,
                    "amount":i["amount"],
                    "request_no":i["request_no"],
                    "module_name":"Freight Payment"
                    })
            else:
                target.append("amount_table",{
                    "benificiary_name":i["beneficiary_name"],
                    "party_type":"Supplier",
                    "party_name":i["beneficiary_name"],
                    "account_no":i["account_no"],
                    "ifsc_code":i["ifsc_code"],
                    "bank_name":i["bank_name"],
                    "branch":i["branch"],
                    "charges":i["charges"],
                    "amount":i["amount"],
                    "request_no":i["request_no"],
                    "module_name":"Freight Payment"
                    })
    doc = get_mapped_doc("Payment Module", source_name, {
        "Payment Module": {
            "doctype": "Final bank statement",
            "validation": {
                "docstatus": ["=", 1]
            }
        },
        "Freight Table":{
            "doctype":"Statement Freight Table"
        }
    }, target_doc,postprocess)
    return doc


@frappe.whitelist()
def make_credit(source_name, target_doc=None):
    def postprocess(source, target):
        item=[]
        for i in source.get('consolidate_table'):
            for j in frappe.db.sql("SELECT bank_account_no,ifsc_code,bank,branch FROM `tabBank Account` WHERE party='"+i.supplier+"'",as_dict=1):
                if(float(i.total)<=float(10000)):
                    crg=2
                elif(float(i.total)>float(10000) and float(i.total)<=float(100000)):
                    crg=5
                elif(float(i.total)>float(100000) and float(i.total)<=float(200000)):
                    crg=13
                elif(float(i.total)>float(200000) and float(i.total)<=float(500000)):
                    crg=28
                elif(float(i.total)>float(500000)):
                    crg=56
                item.append({
                    "beneficiary_name":i.supplier,
                    "account_no":j.bank_account_no,
                    "ifsc_code":j.ifsc_code,
                    "bank_name":j.bank,
                    "branch":j.branch,
                    "charges":crg,
                    "amount":i.total,
                    "parent":i.parent,
                    "request_no":i.name
                    })
        for i in item:
            if(i["bank_name"]=='INDIAN BANK'):
                target.append("ib_amount_table",{
                    "benificiary_name":i["beneficiary_name"],
                    "party_type":"Supplier",
                    "party_name":i["beneficiary_name"],
                    "account_no":i["account_no"],
                    "ifsc_code":i["ifsc_code"],
                    "bank_name":i["bank_name"],
                    "branch":i["branch"],
                    "charges":0,
                    "amount":i["amount"],
                    "request_no":i["request_no"],
                    "payment_module":i["parent"],
                    "module_name":"Invoice Payment"
                    })
            else:
                target.append("amount_table",{
                    "benificiary_name":i["beneficiary_name"],
                    "party_type":"Supplier",
                    "party_name":i["beneficiary_name"],
                    "account_no":i["account_no"],
                    "ifsc_code":i["ifsc_code"],
                    "bank_name":i["bank_name"],
                    "branch":i["branch"],
                    "charges":i["charges"],
                    "amount":i["amount"],
                    "request_no":i["request_no"],
                    "payment_module":i["parent"],
                    "module_name":"Invoice Payment"
                    })
    doc = get_mapped_doc("Payment Module", source_name, {
        "Payment Module": {
            "doctype": "Final bank statement",
            "validation": {
                "docstatus": ["=", 1]
            }
        },
        "Final Overdue Table":{
            "doctype":"Statement Overdue Table"
        }
    }, target_doc,postprocess)
    return doc



@frappe.whitelist()
def make_rop(source_name, target_doc=None):
    def postprocess(source, target):
        item=[]
        for i in frappe.db.sql("SELECT party_type,party,bank_account_no,ifsc_code,bank,branch,grand_total,employee_bank_account_no,employee_bank_name,employee_bank_ifsc,employee_bank_branch FROM `tabRequest for Payment` where name='"+str(source.name)+"' ",as_dict=1):
            if(float(i.grand_total)<=float(10000)):
                crg=2
            elif(float(i.grand_total)>float(10000) and float(i.grand_total)<=float(100000)):
                crg=5
            elif(float(i.grand_total)>float(100000) and float(i.grand_total)<=float(200000)):
                crg=13
            elif(float(i.grand_total)>float(200000) and float(i.grand_total)<=float(500000)):
                crg=28
            elif(float(i.grand_total)>float(500000)):
                crg=56
            if(i.party_type=='Farm Employee'):
                item.append({
                    "beneficiary_name":frappe.get_value("Farm Employee",i.party,"employee_name"),
                    "party_type":i.party_type,
                    "account_no":i.employee_bank_account_no,
                    "ifsc_code":i.employee_bank_ifsc,
                    "bank_name":i.employee_bank_name,
                    "branch":i.employee_bank_branch,
                    "charges":crg,
                    "amount":i.grand_total,
                    "request_no":i.name
                    })
            elif(i.party_type=='Employee'):
                item.append({
                    "beneficiary_name":frappe.get_value("Employee",i.party,"employee_name"),
                    "party_type":i.party_type,
                    "account_no":i.employee_bank_account_no,
                    "ifsc_code":i.employee_bank_ifsc,
                    "bank_name":i.employee_bank_name,
                    "branch":i.employee_bank_branch,
                    "charges":crg,
                    "amount":i.grand_total,
                    "request_no":i.name
                    })
            else:
                item.append({
                    "beneficiary_name":i.party,
                    "party_type":i.party_type,
                    "account_no":i.bank_account_no,
                    "ifsc_code":i.ifsc_code,
                    "bank_name":i.bank,
                    "branch":i.branch,
                    "charges":crg,
                    "amount":i.grand_total,
                    "request_no":i.name
                    })
        for i in item:
            if(i["bank_name"]=='INDIAN BANK'):
                target.append("ib_amount_table",{
                    "benificiary_name":i["beneficiary_name"],
                    "party_type":i["party_type"],
                    "party_name":i["beneficiary_name"],
                    "account_no":i["account_no"],
                    "ifsc_code":i["ifsc_code"],
                    "bank_name":i["bank_name"],
                    "branch":i["branch"],
                    "charges":0,
                    "amount":i["amount"],
                    "request_no":i["request_no"],
                    "module_name":"Request for Payment"
                    })
            else:
                target.append("amount_table",{
                    "benificiary_name":i["beneficiary_name"],
                    "party_type":i["party_type"],
                    "party_name":i["beneficiary_name"],
                    "account_no":i["account_no"],
                    "ifsc_code":i["ifsc_code"],
                    "bank_name":i["bank_name"],
                    "branch":i["branch"],
                    "charges":i["charges"],
                    "amount":i["amount"],
                    "request_no":i["request_no"],
                    "module_name":"Request for Payment"
                    })
    doc = get_mapped_doc("Request for Payment", source_name, {
        "Request for Payment": {
            "doctype": "Final Bank Statement",
            "validation": {
                "docstatus": ["=", 1]
            }
            # "field_map": {
            #     "party": "benificiary_name",
            #     "bank": "bank_name",
            #     "bank_account_no": "account_no",
            #     "ifsc_code": "ifsc_code",
            #     "branch":"branch",
            #     "amount":"grand_total"
            # }
        }
    }, target_doc,postprocess)
    return doc

@frappe.whitelist()
def make_purchase_rop(source_name, target_doc=None):
    def postprocess(source, target):
        item=[]
        for i in frappe.db.sql("SELECT party_type,party,bank_account_no,ifsc_code,bank,branch,total_amount,employee_bank_account_no,employee_bank_name,employee_bank_ifsc,employee_bank_branch FROM `tabPayment Module` where name='"+str(source.name)+"' ",as_dict=1):
            if(float(i.total_amount)<=float(10000)):
                crg=2
            elif(float(i.total_amount)>float(10000) and float(i.total_amount)<=float(100000)):
                crg=5
            elif(float(i.total_amount)>float(100000) and float(i.total_amount)<=float(200000)):
                crg=13
            elif(float(i.total_amount)>float(200000) and float(i.total_amount)<=float(500000)):
                crg=28
            elif(float(i.total_amount)>float(500000)):
                crg=56
            if(i.party_type=='Farm Employee'):
                item.append({
                    "beneficiary_name":frappe.get_value("Farm Employee",i.party,"employee_name"),
                    "party_type":i.party_type,
                    "party_name":i.party,
                    "account_no":i.employee_bank_account_no,
                    "ifsc_code":i.employee_bank_ifsc,
                    "bank_name":i.employee_bank_name,
                    "branch":i.employee_bank_branch,
                    "charges":crg,
                    "amount":i.total_amount,
                    "request_no":i.name
                    })
            elif(i.party_type=='Employee'):
                item.append({
                    "beneficiary_name":frappe.get_value("Employee",i.party,"employee_name"),
                    "party_type":i.party_type,
                    "party_name":i.party,
                    "account_no":i.employee_bank_account_no,
                    "ifsc_code":i.employee_bank_ifsc,
                    "bank_name":i.employee_bank_name,
                    "branch":i.employee_bank_branch,
                    "charges":crg,
                    "amount":i.total_amount,
                    "request_no":i.name
                    })
            else:
                item.append({
                    "beneficiary_name":i.party,
                    "party_type":i.party_type,
                    "account_no":i.bank_account_no,
                    "ifsc_code":i.ifsc_code,
                    "bank_name":i.bank,
                    "branch":i.branch,
                    "charges":crg,
                    "amount":i.total_amount,
                    "request_no":i.name
                    })
        for i in item:
            if(i["bank_name"]=='INDIAN BANK'):
                target.append("ib_amount_table",{
                    "benificiary_name":i["beneficiary_name"],
                    "party_type":i["party_type"],
                    "party_name":i["beneficiary_name"],
                    "account_no":i["account_no"],
                    "ifsc_code":i["ifsc_code"],
                    "bank_name":i["bank_name"],
                    "branch":i["branch"],
                    "charges":0,
                    "amount":i["amount"],
                    "request_no":i["request_no"],
                    "module_name":"Request for Payment for Purchase"
                    })
            else:
                target.append("amount_table",{
                    "benificiary_name":i["beneficiary_name"],
                    "party_type":i["party_type"],
                    "party_name":i["beneficiary_name"],
                    "account_no":i["account_no"],
                    "ifsc_code":i["ifsc_code"],
                    "bank_name":i["bank_name"],
                    "branch":i["branch"],
                    "charges":i["charges"],
                    "amount":i["amount"],
                    "request_no":i["request_no"],
                    "module_name":"Request for Payment for Purchase"
                    })
    doc = get_mapped_doc("Payment Module", source_name, {
        "Payment Module": {
            "doctype": "Final Bank Statement",
            "validation": {
                "docstatus": ["=", 1]
            }
            
        }
    }, target_doc,postprocess)
    return doc


@frappe.whitelist()
def make_lcv(source_name, target_doc=None):
    def postprocess(source, target):
        item=[]
        for i in source.get('landed_table'):
            for j in frappe.db.sql("SELECT bank_account_no,ifsc_code,bank,branch FROM `tabBank Account` WHERE party='"+i.supplier+"'",as_dict=1):
                if(float(i.amount)<=float(10000)):
                    crg=2
                elif(float(i.amount)>float(10000) and float(i.amount)<=float(100000)):
                    crg=5
                elif(float(i.amount)>float(100000) and float(i.amount)<=float(200000)):
                    crg=13
                elif(float(i.amount)>float(200000) and float(i.amount)<=float(500000)):
                    crg=28
                elif(float(i.amount)>float(500000)):
                    crg=56
                item.append({
                    "beneficiary_name":i.supplier,
                    "account_no":j.bank_account_no,
                    "ifsc_code":j.ifsc_code,
                    "bank_name":j.bank,
                    "branch":j.branch,
                    "charges":crg,
                    "amount":i.amount,
                    "request_no":i.name
                    })
        for i in item:
            if(i["bank_name"]=='INDIAN BANK'):
                target.append("ib_amount_table",{
                    "benificiary_name":i["beneficiary_name"],
                    "party_type":"Supplier",
                    "party_name":i["beneficiary_name"],
                    "account_no":i["account_no"],
                    "ifsc_code":i["ifsc_code"],
                    "bank_name":i["bank_name"],
                    "branch":i["branch"],
                    "charges":0,
                    "amount":i["amount"],
                    "request_no":i["request_no"],
                    "module_name":"Invoice Payment"
                    })
            else:
                target.append("amount_table",{
                    "benificiary_name":i["beneficiary_name"],
                    "party_type":"Supplier",
                    "party_name":i["beneficiary_name"],
                    "account_no":i["account_no"],
                    "ifsc_code":i["ifsc_code"],
                    "bank_name":i["bank_name"],
                    "branch":i["branch"],
                    "charges":i["charges"],
                    "amount":i["amount"],
                    "request_no":i["request_no"],
                    "module_name":"Invoice Payment"
                    })
    doc = get_mapped_doc("Payment Module", source_name, {
        "Payment Module": {
            "doctype": "Final bank statement",
            "validation": {
                "docstatus": ["=", 1]
            }
        },
        "Landed Table":{
            "doctype":"Statement LCV"
        }
    }, target_doc,postprocess)
    return doc

@frappe.whitelist()
def make_logistics(source_name, target_doc=None):
    def postprocess(source, target):
        item=[]
        for i in frappe.db.sql("SELECT name,log_supplier,log_rounded_total,log_bank_account,log_bank,log_bank_account_no,log_account,log_branch,log_iban,log_branch_code,log_ifsc_code FROM `tabPayment Module` where name='"+str(source.name)+"' ",as_dict=1):
            if(float(i.log_rounded_total)<=float(10000)):
                crg=2
            elif(float(i.log_rounded_total)>float(10000) and float(i.log_rounded_total)<=float(100000)):
                crg=5
            elif(float(i.log_rounded_total)>float(100000) and float(i.log_rounded_total)<=float(200000)):
                crg=13
            elif(float(i.log_rounded_total)>float(200000) and float(i.log_rounded_total)<=float(500000)):
                crg=28
            elif(float(i.log_rounded_total)>float(500000)):
                crg=56
            item.append({
                "beneficiary_name":i.log_supplier,
                "party_type":"Supplier",
                "account_no":i.log_bank_account_no,
                "ifsc_code":i.log_ifsc_code,
                "bank_name":i.log_bank,
                "branch":i.log_branch,
                "charges":crg,
                "amount":i.log_rounded_total,
                "request_no":i.name
                })
        for i in item:
            if(i["bank_name"]=='INDIAN BANK'):
                target.append("ib_amount_table",{
                    "benificiary_name":i["beneficiary_name"],
                    "party_type":i["party_type"],
                    "party_name":i["beneficiary_name"],
                    "account_no":i["account_no"],
                    "ifsc_code":i["ifsc_code"],
                    "bank_name":i["bank_name"],
                    "branch":i["branch"],
                    "charges":0,
                    "amount":i["amount"],
                    "request_no":i["request_no"],
                    "module_name":"Request for Payment for Purchase"
                    })
            else:
                target.append("amount_table",{
                    "benificiary_name":i["beneficiary_name"],
                    "party_type":i["party_type"],
                    "party_name":i["beneficiary_name"],
                    "account_no":i["account_no"],
                    "ifsc_code":i["ifsc_code"],
                    "bank_name":i["bank_name"],
                    "branch":i["branch"],
                    "charges":i["charges"],
                    "amount":i["amount"],
                    "request_no":i["request_no"],
                    "module_name":"Request for Payment for Purchase"
                    })
    doc = get_mapped_doc("Payment Module", source_name, {
        "Payment Module": {
            "doctype": "Final Bank Statement",
            "validation": {
                "docstatus": ["=", 1]
            }
            
        }
    }, target_doc,postprocess)
    return doc

@frappe.whitelist()
def create_excel(dds):
    to_python = json.loads(dds)
    ur = []
    xl = [["SNO_REF_NO", "CUSTOMER_NAME", "CITY", "ACCOUNT_NO", "AMOUNT", "DESCRIPTION", "IFSC_CODE", "BANK_NAME", "BENEFICIARY_EMAIL_ID"]]
    v = 1

    for i in to_python:
        IFSC_Code = i["ifsc_code"]
        URL = "https://ifsc.razorpay.com/"
        data = requests.get(URL + IFSC_Code).json()

        xl.append([v, i["benificiary_name"], data['CENTRE'], i["account_no"], i["amount"], i["party_type"], i["ifsc_code"], i["bank_name"], "accounts@wttindia.com"])
        v = v + 1

    wb = openpyxl.Workbook()
    ws = wb.active
    for row_data in xl:
        ws.append(row_data)

    sno_ref_no_column = ws['A']
    for cell in sno_ref_no_column:
        cell.alignment = Alignment(horizontal='center')

    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    file_data = openpyxl.writer.excel.save_virtual_workbook(wb)
    file_name = "Sample Bulk RTGS File Format.xlsx"

    _file = frappe.get_doc({
        "doctype": "File",
        "file_name": file_name,
        "folder": "Home/Attachments",
        "content": file_data})
    _file.save()
    ur.append({"url": _file.file_url})
    return ur


@frappe.whitelist()
def create_excel_ib(doc):
    to_python = json.loads(doc)
    ur = []
    xl = [["SNO_REF_NO", "CUSTOMER_NAME", "CITY", "ACCOUNT_NO", "AMOUNT", "DESCRIPTION", "IFSC_CODE", "BANK_NAME", "BENEFICIARY_EMAIL_ID"]]
    v = 1

    for i in to_python:
        IFSC_Code = i["ifsc_code"]
        URL = "https://ifsc.razorpay.com/"
        data = requests.get(URL + IFSC_Code).json()

        xl.append([v, i["benificiary_name"], data['CENTRE'], i["account_no"], i["amount"], i["party_type"], i["ifsc_code"], i["bank_name"], "accounts@wttindia.com"])
        v = v + 1

    wb = openpyxl.Workbook()
    ws = wb.active
    for row_data in xl:
        ws.append(row_data)

    sno_ref_no_column = ws['A']
    for cell in sno_ref_no_column:
        cell.alignment = Alignment(horizontal='center')

    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    file_data = openpyxl.writer.excel.save_virtual_workbook(wb)
    file_name = "Sample Bulk RTGS File Format.xlsx"

    _file = frappe.get_doc({
        "doctype": "File",
        "file_name": file_name,
        "folder": "Home/Attachments",
        "content": file_data})
    _file.save()
    ur.append({"url": _file.file_url})
    return ur

@frappe.whitelist()
def test_func(doc):
    supls=[]
    dd=frappe.get_doc("Final bank statement",doc)
    if(dd.amount_table):
        for i in dd.amount_table:
            supls.append({
                "supplier":i.benificiary_name,
                "amount":i.amount
                })
    if(dd.amount_table):
        for i in dd.ib_amount_table:
            supls.append({
                "supplier":i.benificiary_name,
                "amount":i.amount
                })
    frappe.msgprint(str(supls))